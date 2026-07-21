#!/usr/bin/env python3
"""Pangkalan Geo-Mapper — single-file Python http.server on port 5004"""
import json, sqlite3, os, re
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from pathlib import Path

HOST, PORT = "0.0.0.0", 5004
DB = Path("/root/second-brain/02 - LPG Work/database/pangkalan.db")

CSS = """
:root{--bg:#0f0f1a;--card:#1a1a2e;--card2:#252542;--accent:#7c3aed;--text:#e2e8f0;--muted:#94a3b8;--border:#2d2d44;--danger:#ef4444;--warn:#f59e0b}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:Inter,system-ui,sans-serif;min-height:100vh}
header{background:rgba(26,26,46,.85);backdrop-filter:blur(12px);border-bottom:1px solid var(--border);padding:16px 24px;position:sticky;top:0;z-index:1000}
.header-inner{max-width:1400px;margin:0 auto;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}
header h1{font-size:1.3rem;font-weight:700;display:flex;align-items:center;gap:8px}
header h1 small{font-size:.75rem;color:var(--muted);font-weight:400}
.container{max-width:1400px;margin:0 auto;padding:16px 24px}
.stats{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px}
.stat{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:10px 16px;font-size:.82rem}
.stat strong{font-size:1.2rem;display:block}
.filters{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-bottom:14px;background:var(--card);border:1px solid var(--border);border-radius:12px;padding:14px 16px}
.filters select,.filters input[type=text]{background:var(--bg);border:1px solid var(--border);color:var(--text);padding:8px 12px;border-radius:8px;font-size:.85rem;outline:none}
.filters select:focus,.filters input:focus{border-color:var(--accent)}
.filters label{font-size:.8rem;color:var(--muted);display:flex;align-items:center;gap:6px;cursor:pointer}
.filters label input[type=checkbox]{accent-color:var(--accent);width:16px;height:16px}
.layout{display:grid;grid-template-columns:1fr 400px;gap:16px;min-height:60vh}
@media(max-width:900px){.layout{grid-template-columns:1fr}}
@media(max-width:640px){
header{padding:10px 12px}
header h1{font-size:1rem}
header h1 small{font-size:.65rem}
.container{padding:10px 12px}
.stats{gap:6px}
.stat{padding:7px 10px;font-size:.72rem;min-width:0}
.stat strong{font-size:.95rem;display:inline;margin-right:4px}
#stats{display:grid;grid-template-columns:1fr 1fr;gap:6px}
#stats .stat{margin:0}
.filters{gap:6px;padding:10px 12px;flex-wrap:wrap}
.filters select,.filters input[type=text]{font-size:.78rem;padding:6px 10px;flex:1 1 100%}
.filters label{font-size:.72rem;white-space:nowrap}
.filters button{margin-left:0!important;width:100%;justify-content:center}
#map{min-height:55vh}
.list-wrap{max-height:40vh}
.search-wrap input{font-size:.82rem;padding:8px 12px}
.modal{padding:16px;max-width:100%;margin:10px}
.modal-overlay{padding:10px}
}
.map-wrap{background:var(--card);border:1px solid var(--border);border-radius:12px;overflow:hidden;min-height:500px}
#map{width:100%;height:100%;min-height:500px}
.sidebar{display:flex;flex-direction:column;gap:12px}
.search-wrap input{width:100%;background:var(--card);border:1px solid var(--border);color:var(--text);padding:10px 14px;border-radius:10px;font-size:.9rem;outline:none}
.search-wrap input:focus{border-color:var(--accent)}
.list-wrap{background:var(--card);border:1px solid var(--border);border-radius:12px;flex:1;overflow-y:auto;max-height:600px}
.list-item{padding:10px 14px;border-bottom:1px solid var(--border);cursor:pointer;transition:background .1s}
.list-item:hover{background:var(--card2)}
.list-item .name{font-weight:500;font-size:.9rem}
.list-item .meta{font-size:.78rem;color:var(--muted);margin-top:3px;display:flex;gap:8px;flex-wrap:wrap}
.badge{font-size:.7rem;padding:2px 8px;border-radius:99px;font-weight:600}
.badge-active{background:rgba(34,197,94,.15);color:#22c55e}
.badge-inactive{background:rgba(239,68,68,.15);color:#ef4444}
.badge-problem{background:rgba(245,158,11,.15);color:var(--warn)}
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.7);backdrop-filter:blur(4px);z-index:9999;display:none;align-items:center;justify-content:center;padding:20px}
.modal-overlay.active{display:flex}
.modal{background:var(--card);border:1px solid var(--border);border-radius:14px;width:100%;max-width:480px;max-height:80vh;overflow-y:auto;padding:24px;box-shadow:0 20px 60px rgba(0,0,0,.5)}
.modal h2{font-size:1.15rem;margin-bottom:16px}
.modal .field{margin-bottom:10px}
.modal .field label{font-size:.75rem;color:var(--muted);display:block;margin-bottom:2px}
.modal .field .val{font-size:.9rem}
.modal-close{float:right;background:none;border:none;color:var(--muted);font-size:1.5rem;cursor:pointer}
.modal-close:hover{color:var(--text)}
footer{text-align:center;color:var(--muted);font-size:.8rem;padding:20px;border-top:1px solid var(--border);margin-top:20px}
.leaflet-popup-content-wrapper{background:var(--card)!important;color:var(--text)!important;border-radius:10px!important}
.leaflet-popup-tip{background:var(--card)!important}
"""

HTML_TEMPLATE = """<!doctype html>
<html lang="id"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Pangkalan Geo-Mapper</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<style>__CSS__</style></head><body>
<header><div class="header-inner">
<h1>&#x1F5FA;&#xFE0F; Pangkalan Geo-Mapper <small>Kalimantan Barat</small></h1>
<div id="timeBar" style="font-size:.8rem;color:var(--muted)"></div>
</div></header>
<div class="container">
<div class="stats" id="stats"></div>
<div class="filters" id="filters">
<select id="filterStatus"><option value="Active">Active</option><option value="">Semua Status</option><option value="Inactive">Inactive</option></select>
<select id="filterKota" style="min-width:180px"><option value="">Semua Kota</option></select>
<input type="text" id="filterSearch" placeholder="Cari nama pangkalan..." style="flex:1;min-width:140px">
<label><input type="checkbox" id="probQty"> Qty > 3000</label>
<label><input type="checkbox" id="probKota"> Tanpa Kota</label>
<label><input type="checkbox" id="probNib"> NIB kosong</label>

<button onclick="exportExcel()" style="margin-left:auto;background:var(--accent);color:#fff;border:none;padding:8px 16px;border-radius:8px;cursor:pointer;font-size:.85rem;font-weight:600">📥 Export Excel</button>
</div>
<div class="layout">
<div class="map-wrap"><div id="map"></div></div>
<div class="sidebar"><div class="search-wrap"><input id="listSearch" placeholder="Cari di daftar..." oninput="filterList()"></div>
<div class="list-wrap" id="listWrap"></div></div></div></div>
<div class="modal-overlay" id="modalOverlay"><div class="modal" id="modalContent"></div></div>
<footer>Pangkalan Geo-Mapper &mdash; Data: pangkalan.db (5.299 records)</footer>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js">
function exportExcel(){
 var s=document.getElementById("filterStatus").value, k=document.getElementById("filterKota").value, q=document.getElementById("filterSearch").value;
 var prob=[];
 if(document.getElementById("probQty").checked)prob.push("qty");
 if(document.getElementById("probKota").checked)prob.push("kota");
 if(document.getElementById("probNib").checked)prob.push("nib");
 
 var p=prob.length?"&problem="+prob.join(","):"";
 window.open("__BASE__/api/pangkalan/export?status="+s+"&kota="+k+"&q="+q+p, "_blank");
}
</script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<script>
const map = L.map('map').setView([-0.02, 109.35], 10);
L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png',{maxZoom:19}).addTo(map);
const markers = L.markerClusterGroup({chunkedLoading:true,maxClusterRadius:50,spiderfyOnMaxZoom:true});
map.addLayer(markers);
let allData = [];

function load(){
 const s=document.getElementById('filterStatus').value, k=document.getElementById('filterKota').value, q=document.getElementById('filterSearch').value;
 const prob=[];
 if(document.getElementById('probQty').checked)prob.push('qty');
 if(document.getElementById('probKota').checked)prob.push('kota');
 if(document.getElementById('probNib').checked)prob.push('nib');
 const p=prob.length?'&problem='+prob.join(','):'';
 fetch('__BASE__/api/pangkalan?status='+s+'&kota='+k+'&q='+q+p).then(r=>r.json()).then(d=>{
  allData=d.data||[]; renderMap(); }).catch(e=>{
  document.getElementById('stats').innerHTML='<div class="stat" style="border-color:var(--danger)">Error: '+e.message+'</div>';})}

function renderMap(){
 markers.clearLayers();
 const st=document.getElementById('stats');
 const active=allData.filter(x=>x.status==='Active').length;
 const probCount=allData.filter(x=>x.problems&&x.problems.length).length;
 st.innerHTML='<div class="stat"><strong>'+allData.length+'</strong> Pangkalan</div>'+
  '<div class="stat"><strong>'+active+'</strong> Active</div>'+
  '<div class="stat"><strong>'+(allData.length-active)+'</strong> Inactive</div>';
 if(probCount)st.innerHTML+='<div class="stat" style="border-color:var(--warn)"><strong>'+probCount+'</strong> &#x26A0;&#xFE0F; Bermasalah</div>';
 renderList();
 allData.forEach(function(p){
  if(!p.lat||!p.lng||p.lat==0||p.lng==0)return;
  var pb=p.problems&&p.problems.length?'<br><span style="color:var(--warn)">&#x26A0;&#xFE0F; '+p.problems.join(', ')+'</span>':'';
  var m=L.marker([p.lat,p.lng]);
  m.on('click',function(){showDetail(p.id_registrasi);});
  markers.addLayer(m);})}

function renderList(){
 var q=(document.getElementById('listSearch').value||'').toLowerCase();
 var items=allData.filter(function(x){return x.nama_pangkalan&&x.nama_pangkalan.toLowerCase().includes(q);});
 document.getElementById('listWrap').innerHTML=items.map(function(p){
  var cls=p.status==='Active'?'badge-active':'badge-inactive';
  var prob=p.problems&&p.problems.length?'<span class="badge badge-problem">&#x26A0;&#xFE0F;</span>':'';
  return '<div class="list-item" onclick="showDetail(\\''+p.id_registrasi+'\\')"><div class="name">'+(p.nama_pangkalan||'?')+' '+prob+'</div><div class="meta"><span class="badge '+cls+'">'+(p.status||'?')+'</span><span>'+(p.kota||'-')+'</span><span>Qty: '+(p.qty_kontrak||0)+'</span></div></div>';
 }).join('');}

function showDetail(id){
 fetch('__BASE__/api/pangkalan/detail?id_registrasi='+id).then(function(r){return r.json();}).then(function(p){
  if(p.error)return;
  var pb=p.problems&&p.problems.length?'<div style="color:var(--warn);margin-top:8px">&#x26A0;&#xFE0F; '+p.problems.join(', ')+'</div>':'';
  document.getElementById('modalContent').innerHTML='<button class="modal-close" onclick="closeModal()">&times;</button>'+
  '<h2>'+(p.nama_pangkalan||'-')+'</h2>'+
  '<div class="field"><label>Pemilik</label><div class="val">'+(p.nama_pemilik||'-')+'</div></div>'+
  '<div class="field"><label>Alamat</label><div class="val">'+(p.alamat||'-')+', '+(p.kelurahan||'')+', '+(p.kecamatan||'')+', '+(p.kota||'')+'</div></div>'+
  '<div class="field"><label>Kota</label><div class="val">'+(p.kota||'-')+'</div></div>'+
  '<div class="field"><label>Kecamatan</label><div class="val">'+(p.kecamatan||'-')+'</div></div>'+
  '<div class="field"><label>Status</label><div class="val"><span class="badge '+(p.status==='Active'?'badge-active':'badge-inactive')+'">'+(p.status||'-')+'</span></div></div>'+
  '<div class="field"><label>Qty Kontrak</label><div class="val">'+(p.qty_kontrak||0)+'</div></div>'+
  '<div class="field"><label>ID Registrasi</label><div class="val">'+(p.id_registrasi||'-')+'</div></div>'+
  '<div class="field"><label>No. HP</label><div class="val">'+(p.no_hp_pemilik||'-')+'</div></div>'+
  '<div class="field"><label>NIB</label><div class="val">'+(p.nomor_nib||'<span style="color:var(--warn)">Kosong</span>')+'</div></div>'+
  '<div class="field"><label>Integrasi Map</label><div class="val">'+(p.integrasi_map||'-')+'</div></div>'+
  '<div class="field"><label>Koordinat</label><div class="val">'+(p.lat&&p.lat!=0?p.lat+', '+p.lng:'<span style="color:var(--warn)">Tidak ada</span>')+'</div></div>'+
  pb+
  '<div style="margin-top:16px"><button class="modal-close" style="float:none;background:var(--accent);color:#fff;border:none;padding:8px 20px;border-radius:8px;cursor:pointer;font-size:.9rem" onclick="closeModal()">Tutup</button></div>';
  document.getElementById('modalOverlay').classList.add('active');});}

function closeModal(){document.getElementById('modalOverlay').classList.remove('active');}
document.getElementById('modalOverlay').onclick=function(e){if(e.target===document.getElementById('modalOverlay'))closeModal();};
function filterList(){renderList();}

fetch('__BASE__/api/pangkalan/kota').then(function(r){return r.json();}).then(function(d){
 var sel=document.getElementById('filterKota');
 sel.innerHTML='<option value="">Semua Kota</option>'+d.map(function(k){return '<option value="'+k+'">'+k+'</option>';}).join('');
 sel.value='KOTA PONTIANAK';
}).then(function(){load();});

document.getElementById('filterStatus').onchange=load;
document.getElementById('filterKota').onchange=load;
document.getElementById('filterSearch').oninput=load;
document.querySelectorAll('#filters input[type=checkbox]').forEach(function(c){c.onchange=load;});

function updateTime(){
 document.getElementById('timeBar').textContent='🕐 '+new Date().toLocaleDateString('id-ID',{weekday:'long',year:'numeric',month:'long',day:'numeric',hour:'2-digit',minute:'2-digit',second:'2-digit',timeZone:'Asia/Pontianak'});}
updateTime();setInterval(updateTime,1000);
</script></body></html>"""

HTML = HTML_TEMPLATE.replace("__CSS__", CSS).replace("__BASE__", "/pangkalan")

class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path.rstrip("/") or "/"
        params = parse_qs(parsed.query)
        
        # Strip /pangkalan prefix if present (for nginx sub-path deployment)
        if path.startswith("/pangkalan"):
            path = path[len("/pangkalan"):] or "/"
        
        if path == "/":
            self._html(HTML)
            return
        
        if path == "/api/pangkalan/kota":
            self._json(self._get_kota())
            return
        if path == "/api/pangkalan/stats":
            self._json(self._get_stats())
            return
        if path == "/api/pangkalan/detail":
            id_reg = params.get("id_registrasi", [""])[0]
            self._json(self._get_detail(id_reg))
            return
        if path == "/api/pangkalan":
            data = self._get_pangkalan(params)
            self._json(data)
            return
        if path == "/api/pangkalan/export":
            self._export_excel(params)
            return
        
        self._error("Not found", 404)
    
    def _db(self):
        conn = sqlite3.connect(str(DB))
        conn.row_factory = sqlite3.Row
        return conn
    
    def _get_kota(self):
        try:
            conn = self._db()
            rows = conn.execute("SELECT DISTINCT kota FROM pangkalan WHERE kota IS NOT NULL AND kota != '' ORDER BY kota").fetchall()
            conn.close()
            return [r["kota"] for r in rows if r["kota"]]
        except: return []
    
    def _get_stats(self):
        try:
            conn = self._db()
            r = conn.execute("SELECT COUNT(*) as t, SUM(CASE WHEN status='Active' THEN 1 ELSE 0 END) as a FROM pangkalan").fetchone()
            prob_qty = conn.execute("SELECT COUNT(*) FROM pangkalan WHERE qty_kontrak > 3000").fetchone()[0]
            prob_kota = conn.execute("SELECT COUNT(*) FROM pangkalan WHERE kota IS NULL OR kota=''").fetchone()[0]
            prob_nib = conn.execute("SELECT COUNT(*) FROM pangkalan WHERE nomor_nib IS NULL OR nomor_nib=''").fetchone()[0]
            prob_koord = conn.execute("SELECT COUNT(*) FROM pangkalan WHERE latitude IS NULL OR latitude=0").fetchone()[0]
            conn.close()
            return {"total":r["t"],"active":r["a"],"inactive":r["t"]-r["a"],
                    "problems":{"qty":prob_qty,"kota":prob_kota,"nib":prob_nib,"koordinat":prob_koord}}
        except: return {}
    
    def _get_detail(self, id_reg):
        if not id_reg: return {"error":"No ID"}
        try:
            conn = self._db()
            r = conn.execute("SELECT * FROM pangkalan WHERE id_registrasi=?", (id_reg,)).fetchone()
            conn.close()
            if not r: return {"error":"Not found"}
            d = dict(r)
            d["problems"] = self._check_problems(d)
            d["lat"] = d.get("latitude")
            d["lng"] = d.get("longitude")
            return d
        except Exception as e: return {"error":str(e)}
    
    def _check_problems(self, d):
        problems = []
        try:
            if d.get("qty_kontrak") and int(d["qty_kontrak"]) > 3000: problems.append("Qty > 3000")
        except: pass
        if not d.get("kota"): problems.append("Tidak ada kota")
        if not d.get("nomor_nib"): problems.append("NIB kosong")
        try:
            lat = float(d.get("latitude", 0) or 0)
            if lat == 0: problems.append("No koordinat")
        except: problems.append("No koordinat")
        return problems
    
    def _get_pangkalan(self, params):
        status = params.get("status", [""])[0]
        kota = params.get("kota", [""])[0]
        search = params.get("q", [""])[0]
        problem = params.get("problem", [""])[0]
        
        where = ["1=1"]
        args = []
        if status == "Active": where.append("status='Active'")
        elif status == "Inactive": where.append("(status='Inactive' OR status IS NULL OR status='')")
        if kota: where.append("kota=?"); args.append(kota)
        if search: where.append("nama_pangkalan LIKE ?"); args.append(f"%{search}%")
        
        prob_clauses = []
        if "qty" in problem: prob_clauses.append("qty_kontrak>3000")
        if "kota" in problem: prob_clauses.append("(kota IS NULL OR kota='')")
        if "nib" in problem: prob_clauses.append("(nomor_nib IS NULL OR nomor_nib='')")
        if prob_clauses: where.append("(" + " OR ".join(prob_clauses) + ")")
        
        query = f"SELECT id_registrasi,nama_pangkalan,kota,qty_kontrak,latitude,longitude,status,nomor_nib FROM pangkalan WHERE {' AND '.join(where)} ORDER BY nama_pangkalan"
        try:
            conn = self._db()
            rows = conn.execute(query, args).fetchall()
            conn.close()
            data = []
            for r in rows:
                d = dict(r)
                d["lat"] = d.get("latitude")
                d["lng"] = d.get("longitude")
                d["problems"] = self._check_problems(d)
                data.append(d)
            return {"data":data,"total":len(data)}
        except Exception as e: return {"error":str(e),"data":[]}
    
    def _export_excel(self, params):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
            import datetime
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pangkalan"
            
            # Header style
            hdr_font = Font(name="Inter", bold=True, color="FFFFFF", size=11)
            hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="2d2d44"),
                right=Side(style="thin", color="2d2d44"),
                top=Side(style="thin", color="2d2d44"),
                bottom=Side(style="thin", color="2d2d44"),
            )
            
            # Fetch data with same filters
            data = self._get_pangkalan(params)
            rows = data.get("data", [])
            
            # Headers
            headers = ["ID Registrasi", "Nama Pangkalan", "Kota", "Kecamatan", "Kelurahan",
                       "Alamat", "Pemilik", "No. HP", "NIB", "Status", "Qty Kontrak",
                       "Latitude", "Longitude", "Integrasi Map", "Masalah"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = hdr_align
                cell.border = thin_border
            
            # Data rows - re-query with full detail
            where = ["1=1"]
            args = []
            s = params.get("status", [""])[0]
            kota = params.get("kota", [""])[0]
            search = params.get("q", [""])[0]
            problem = params.get("problem", [""])[0]
            
            if s == "Active": where.append("status='Active'")
            elif s == "Inactive": where.append("(status='Inactive' OR status IS NULL OR status='')")
            if kota: where.append("kota=?"); args.append(kota)
            if search: where.append("nama_pangkalan LIKE ?"); args.append(f"%{search}%")
            
            prob_clauses = []
            if "qty" in problem: prob_clauses.append("qty_kontrak>3000")
            if "kota" in problem: prob_clauses.append("(kota IS NULL OR kota='')")
            if "nib" in problem: prob_clauses.append("(nomor_nib IS NULL OR nomor_nib='')")
            if prob_clauses: where.append("(" + " OR ".join(prob_clauses) + ")")
            
            query = f"SELECT * FROM pangkalan WHERE {' AND '.join(where)} ORDER BY nama_pangkalan"
            conn = self._db()
            db_rows = conn.execute(query, args).fetchall()
            conn.close()
            
            data_font = Font(name="Inter", size=10)
            data_align = Alignment(vertical="center")
            
            for i, row in enumerate(db_rows, 2):
                d = dict(row)
                problems = self._check_problems(d)
                prob_str = ", ".join(problems) if problems else ""
                vals = [
                    d.get("id_registrasi",""), d.get("nama_pangkalan",""),
                    d.get("kota",""), d.get("kecamatan",""), d.get("kelurahan",""),
                    d.get("alamat",""), d.get("nama_pemilik",""), d.get("no_hp_pemilik",""),
                    d.get("nomor_nib",""), d.get("status",""), d.get("qty_kontrak",0),
                    d.get("latitude",""), d.get("longitude",""), d.get("integrasi_map",""),
                    prob_str
                ]
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=i, column=col, value=v)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
            
            # Auto-width columns
            for col in range(1, len(headers)+1):
                max_len = len(str(headers[col-1]))
                for row in ws.iter_rows(min_col=col, max_col=col, min_row=2, max_row=min(len(db_rows)+1, 50)):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 3, 45)
            
            # Save to BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            content = buf.getvalue()
            
            filename = f"pangkalan_{datetime.date.today().isoformat()}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f"attachment; filename={filename}")
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        except Exception as e:
            self._error(f"Export gagal: {str(e)}", 500)
    
    def _html(self, content, status=200):
        body = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type","text/html; charset=utf-8")
        self.send_header("Content-Length",str(len(body)))
        self.end_headers()
        self.wfile.write(body)
    
    def _json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type","application/json; charset=utf-8")
        self.send_header("Content-Length",str(len(body)))
        self.send_header("Access-Control-Allow-Origin","*")
        self.end_headers()
        self.wfile.write(body)
    
    def _error(self, msg, status=400):
        self._json({"error":msg}, status)
    
    def log_message(self, fmt, *args):
        print(f"[{self.log_date_time_string()}] {args[0]} {args[1]} {args[2]}")

if __name__ == "__main__":
    HTTPServer((HOST, PORT), Handler).serve_forever()
