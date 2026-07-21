#!/usr/bin/env python3
"""Analyze the BBM Excel data"""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('/root/.hermes/cache/documents/doc_49ca0b261285_Laporan_Pembongkaran_dan_Penerimaan_BBM_di_SPBU_MTD_14_Juli_2026.xlsx')
ws = wb['Form Responses 1']

bbm_counter = Counter()
kota_counter = Counter()
spbu_data = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    ts, nama_pt, nomor_spbu, kota, *rest = row
    jenis_bbm = rest[-1] if rest else ''
    
    kota = str(kota or '').strip().title()
    nama = str(nama_pt or '').strip().upper()
    bbm = str(jenis_bbm or '').strip()
    
    if kota:
        kota_counter[kota] += 1
    if nama:
        spbu_data[nama] = spbu_data.get(nama, {'count': 0, 'kota': kota})
        spbu_data[nama]['count'] += 1
    if bbm:
        for item in bbm.split(','):
            bbm_counter[item.strip()] += 1

print("=== RINGKASAN DATA ===")
print(f"Total entries: {ws.max_row - 1:,}")
print(f"Date range: {ws.cell(2,1).value}  →  {ws.cell(ws.max_row,1).value}")
print()
print("=== KOTA / KABUPATEN ===")
for k, v in kota_counter.most_common():
    print(f"  {k}: {v} laporan")
print()
print("=== JENIS BBM ===")
for k, v in bbm_counter.most_common():
    print(f"  {k}: {v}")
print()
print(f"=== SPBU UNIK: {len(spbu_data)} ===")
for nama, info in sorted(spbu_data.items(), key=lambda x: -x[1]['count'])[:15]:
    print(f"  {nama:40s} | {info['kota']:15s} | {info['count']} laporan")
