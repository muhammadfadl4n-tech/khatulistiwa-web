#!/usr/bin/env python3
"""Check which Excel photo links are missing - using psql directly."""
import openpyxl, re, subprocess, sys
from datetime import datetime
from collections import defaultdict

EXCEL = '/root/.hermes/cache/documents/doc_82239650f0f4_Laporan_Pembongkaran_dan_Penerimaan_BBM_di_SPBU_MTD_14_Juli_2026.xlsx'
wb = openpyxl.load_workbook(EXCEL)
ws = wb['Form Responses 1']
categories = ['pembongkaran', 'spp', 'dipping', 'atg']
total_rows = ws.max_row - 1

# Build lookup: (nomor_spbu, tanggal_iso) -> list of (cat, file_id_partial)
# from PostgreSQL
def psql(sql):
    result = subprocess.run(
        ['sudo', '-u', 'postgres', 'psql', '-d', 'qqspbu', '-tA', '-F', '|', '-c', sql],
        capture_output=True, text=True, timeout=30
    )
    return [line.strip() for line in result.stdout.strip().split('\n') if line.strip()]

# Get all uploads
upload_rows = psql("""
    SELECT u.id_laporan, l.id_spbu, s.nomor_spbu, l.tanggal::text, u.kategori, u.path
    FROM uploads u
    JOIN laporan l ON u.id_laporan = l.id
    JOIN spbu s ON l.id_spbu = s.id
""")
print(f"Upload records in DB: {len(upload_rows)}", file=sys.stderr)

# Build lookup set: (nomor_spbu, tanggal, kategori, file_id_prefix)
existing = set()
for row in upload_rows:
    if '|' not in row:
        continue
    parts = row.split('|')
    if len(parts) >= 6:
        nomor = parts[2].strip()
        tanggal = parts[3].strip()
        kategori = parts[4].strip()
        path = parts[5].strip()
        # Extract file_id from path (format: laporan_id/kategori_fileid10.jpg)
        file_id_part = path.split('_')[-1].replace('.jpg','') if '_' in path else ''
        existing.add((nomor, tanggal, kategori, file_id_part[:10]))

print(f"Existing upload tuples: {len(existing)}", file=sys.stderr)

# Also get all laporan to map (nomor, tanggal) -> exists
laporan_set = set()
laporan_rows = psql("""
    SELECT s.nomor_spbu, l.tanggal::text
    FROM laporan l JOIN spbu s ON l.id_spbu = s.id
""")
for row in laporan_rows:
    if '|' not in row:
        continue
    parts = row.split('|')
    if len(parts) >= 2:
        laporan_set.add((parts[0].strip(), parts[1].strip()))

print(f"Laporan records: {len(laporan_set)}", file=sys.stderr)

# Count totals from Excel
cat_total = defaultdict(int)
for row in ws.iter_rows(min_row=2, values_only=True):
    ts, nama_pt, nomor_spbu, kota, *rest = row
    for ci, cat in enumerate(categories):
        url_str = str(rest[ci] or '') if ci < len(rest) else ''
        urls = [u.strip() for u in re.split(r'[,\n\r]+', url_str) if u.strip() and 'drive' in u.lower()]
        cat_total[cat] += len(urls)

total_all_links = sum(cat_total.values())

all_missing = 0
missing_by_cat = defaultdict(int)
missing_by_spbu = defaultdict(lambda: {'count': 0, 'kota': '', 'nama': ''})
rows_missing = 0

for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    ts, nama_pt, nomor_spbu, kota, *rest = row
    nomor = str(int(nomor_spbu)) if isinstance(nomor_spbu, (int, float)) else str(nomor_spbu or '').strip()
    tanggal_obj = ts.date() if isinstance(ts, datetime) else None
    if not tanggal_obj:
        continue
    
    tanggal_str = tanggal_obj.isoformat()
    row_missing = False
    
    for ci, cat in enumerate(categories):
        url_str = str(rest[ci] or '') if ci < len(rest) else ''
        urls = [u.strip() for u in re.split(r'[,\n\r]+', url_str) if u.strip() and 'drive' in u.lower()]
        
        for single_url in urls:
            file_id_match = re.search(r'id=([a-zA-Z0-9_-]+)', single_url)
            if not file_id_match:
                all_missing += 1
                missing_by_cat[cat] += 1
                continue
            
            file_id = file_id_match.group(1)
            file_prefix = file_id[:10]

            # Check if this file_id already uploaded
            found = False
            for existing_nomor, existing_tgl, existing_cat, existing_prefix in existing:
                if existing_nomor == nomor and existing_tgl == tanggal_str and existing_cat == cat and existing_prefix == file_prefix:
                    found = True
                    break
            
            if not found:
                all_missing += 1
                missing_by_cat[cat] += 1
                row_missing = True
                n = nomor
                if n not in missing_by_spbu:
                    missing_by_spbu[n] = {'count': 0, 'kota': str(kota or '').strip().title(), 'nama': str(nama_pt or '').strip().upper()}
                missing_by_spbu[n]['count'] += 1
    
    if row_missing:
        rows_missing += 1

done_count = total_all_links - all_missing

print()
print(f"Total rows in Excel: {total_rows}")
print()
print("=== STATUS FOTO ===")
print(f"Total link di Excel:    {total_all_links}")
print(f"✅ Sudah terupload:     {done_count} ({done_count/total_all_links*100:.1f}%)" if total_all_links else "✅ Tidak ada link foto di Excel")
print(f"❌ BELUM terupload:     {all_missing} ({all_missing/total_all_links*100:.1f}%)" if total_all_links else "")
print(f"Baris kurang foto:      {rows_missing} dari {total_rows}")
print()

print("=== PER KATEGORI FOTO ===")
print(f"  {'Kategori':15s} {'Progress':>20s}")
print(f"  {'-'*35}")
for cat in categories:
    t = cat_total.get(cat, 0)
    m = missing_by_cat.get(cat, 0)
    d = t - m
    pct = d/t*100 if t > 0 else 0
    bar = int(d/t*20) if t > 0 else 0
    print(f"  {cat:15s} {d:4d}/{t:4d} ({pct:5.1f}%) {'█' * bar}")

print()
if missing_by_spbu:
    print("=== TOP 15 SPBU DENGAN FOTO PALING BANYAK KURANG ===")
    print(f"  {'No. SPBU':12s} {'Nama PT':35s} {'Kota':15s} {'Kurang':>8s}")
    print(f"  {'-'*72}")
    sorted_spbu = sorted(missing_by_spbu.items(), key=lambda x: -x[1]['count'])
    for nomor, data in sorted_spbu[:15]:
        print(f"  {nomor:12s} {data['nama']:35s} {data['kota']:15s} {data['count']:8d}")
    print(f"\nTotal SPBU dengan foto kurang: {len(missing_by_spbu)}")
else:
    print("✅ Semua foto sudah terupload dengan sempurna!")
