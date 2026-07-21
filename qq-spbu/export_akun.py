#!/usr/bin/env python3
"""Export all SPBU akun ke Excel"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
os.environ['FLASK_ENV'] = 'production'
os.environ['DATABASE_URL'] = 'postgresql://qqspbu:qqspbu2026@localhost:5432/qqspbu'
os.environ['SECRET_KEY'] = 'qq-spbu-secret-key-2026-kalbar'

from app import app
from models import db, SPBU, User
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = os.path.join(os.path.dirname(__file__), 'daftar_akun_spbu.xlsx')

with app.app_context():
    spbu_list = SPBU.query.order_by(SPBU.kota, SPBU.nama_pt).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Akun SPBU"
    
    # Styles
    hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    b = Border(left=Side(style='thin'), right=Side(style='thin'),
               top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['No', 'Nomor SPBU', 'Nama PT / CV', 'Kota', 'Username', 'Password', 'URL Login']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, b
    
    for i, spbu in enumerate(spbu_list, 1):
        user = User.query.filter_by(id_spbu=spbu.id, role='spbu').first()
        username = user.username if user else '-'
        password = f"spbu{spbu.nomor_spbu[-4:]}" if len(spbu.nomor_spbu) >= 4 else 'spbu123'
        
        row_data = [i, spbu.nomor_spbu, spbu.nama_pt, spbu.kota, username, password, 'https://qq-spbu.khatulistiwa.cloud']
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=c, value=val)
            cell.border = b
            if c == 1: cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    for i, w in enumerate([6, 16, 35, 18, 22, 16, 38], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'
    
    # Add notes sheet
    ws2 = wb.create_sheet("Info")
    notes = [
        ["DAFTAR AKUN SPBU - QQ SPBU"],
        [""],
        ["Total akun:", len(spbu_list)],
        ["URL Login:", "https://qq-spbu.khatulistiwa.cloud"],
        [""],
        ["PENTING:"],
        ["- Password bisa direset oleh Admin di menu Admin > Users"],
        ["- Setiap SPBU hanya bisa lihat laporan miliknya sendiri"],
        ["- Username tidak bisa diubah"],
    ]
    for ri, row in enumerate(notes, 1):
        cell = ws2.cell(row=ri, column=1, value=row[0] if row else '')
        if ri == 1: cell.font = Font(bold=True, size=14, color='1E40AF')
        if len(row) > 1:
            ws2.cell(row=ri, column=2, value=row[1])
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    
    wb.save(OUTPUT)
    print(f"✅ Created: {OUTPUT}")
    print(f"   Total: {len(spbu_list)} akun SPBU")
