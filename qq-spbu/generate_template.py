#!/usr/bin/env python3
"""Generate Excel template untuk import SPBU + Akun User"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

BASE = os.path.dirname(__file__)

def create_styles():
    hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ef = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    b = Border(left=Side(style='thin'), right=Side(style='thin'),
               top=Side(style='thin'), bottom=Side(style='thin'))
    return hf, hfill, ha, ef, b

def create(filename, with_example=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data SPBU + Akun"
    hf, hfill, ha, ef, b = create_styles()

    headers = ['Nomor SPBU','Nama PT / CV','Kota','Alamat','Latitude','Longitude',
               'Username (login)','Password','Nama Lengkap User']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, b

    if with_example:
        ex = ['64.751.01','PT. BBM Sejahtera','Pontianak','Jl. Gajah Mada No. 123',
              '-0.0263','109.3425','spbu01','spbu123','Ahmad Rizki']
        for c, v in enumerate(ex, 1):
            cell = ws.cell(row=2, column=c, value=v)
            cell.fill, cell.border = ef, b

    for i, w in enumerate([16,30,18,40,12,12,16,16,25], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet("Petunjuk")
    notes = [
        ["PETUNJUK PENGISIAN TEMPLATE AKUN SPBU"],
        [""],
        ["1. Isi data pada sheet 'Data SPBU + Akun'"],
        ["2. Kolom wajib: Nomor SPBU, Nama PT, Kota, Username, Password, Nama Lengkap"],
        ["3. Latitude/Longitude opsional (format desimal)"],
        ["4. Username harus unik"],
        ["5. Password minimal 6 karakter"],
        ["6. Setelah diisi, jalankan:  python3 import_akun_spbu.py"],
        [""],
        ["KOLOM:", "Nomor SPBU | Nama PT | Kota | Alamat | Lat | Lng | Username | Password | Nama Lengkap"],
    ]
    for ri, row in enumerate(notes, 1):
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color='1E40AF')
    ws2.column_dimensions['A'].width = 60

    wb.save(filename)
    print(f"Created: {filename}")

create(os.path.join(BASE, "template_akun_spbu.xlsx"), False)
create(os.path.join(BASE, "contoh_isi_akun_spbu.xlsx"), True)
print("Done!")
