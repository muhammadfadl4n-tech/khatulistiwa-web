#!/usr/bin/env python3
"""
Import massal data QQ SPBU:
1. SPBU (nama dibenerin) + buat akun user
2. Laporan
3. Download foto dari Google Drive
"""
import os, sys, re, time, json, subprocess, io
sys.path.insert(0, os.path.dirname(__file__))
os.environ['FLASK_ENV'] = 'production'
os.environ['DATABASE_URL'] = 'postgresql://qqspbu:qqspbu2026@localhost:5432/qqspbu'
os.environ['SECRET_KEY'] = 'qq-spbu-secret-key-2026-kalbar'
from datetime import datetime
from collections import OrderedDict, defaultdict
import openpyxl
from app import app
from models import db, SPBU, User, Laporan, Upload
from utils import compress_image, generate_thumbnail

EXCEL = '/root/.hermes/cache/documents/doc_82239650f0f4_Laporan_Pembongkaran_dan_Penerimaan_BBM_di_SPBU_MTD_14_Juli_2026.xlsx'
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')

# ============================================================
# NAMA SPBU YANG SUDAH DIBENERIN
# ============================================================
NAMA_BENAR = {
    '6178101': 'PT PERTAMINA RETAIL', '6178121': 'PT HEMAT SEJATAR BERSAMA',
    '6178301': 'PT PERTAMINA RETAIL', '63781002': 'PT SOLA PETRA ENERGI',
    '6378101': 'PT BERKAH TOHA', '6378301': 'PT SOLA PETRA ABADI',
    '63783020': 'PT HAYKAL LESTARI', '6468507': 'PT JOHNDITOR',
    '6476812': 'PT ANEKA CIPTA LESTARI', '64778114': 'PT PALEM INDAH LESTARI',
    '6478101': 'PT BERINGIN RAYA', '6478102': 'PT OSO KOTRINDO',
    '6478103': 'PT BERINGIN RAYA', '6478104': 'PT INTI BUANA NUSANTARA',
    '6478105': 'PT MERCAM PONTI BERSATU', '6478106': 'PT MERITA ABADI SUKSES',
    '6478108': 'PT PAWAN RAYA UTAMA', '6478109': 'PT SYENNI BORNEO',
    '6478110': 'PT HEMAD SEJAHTERA BERSAMA', '6478111': 'PT SOLA PETRA ABADI',
    '6478112': 'PT PERMATA PANCASILA', '6478114': 'PT PALEM INDAH LESTARI',
    '6478115': 'PT RACHMAT TIGA PUTRA', '6478116': 'PT HILNASRI PUTRA KALIMANTAN',
    '6478118': 'PT MEGAH BORNEO', '6478119': 'PT CITRA PUTRA MANDIRI',
    '6478120': 'PT BAHRAIN PETRO KHATULISTIWA', '6478121': 'PT HEMAD SEJAHTERA',
    '6478123': 'PT DUTA DAYA KALIMANTAN', '6478201': 'PT BATU LAYANG JAYA',
    '6478202': 'PT BERINGIN RAYA', '6478203': 'PT JOVA JOYA',
    '6478301': 'PT LIONG PUTRA JAYA', '64783020': 'PT HAYKAL LESTARI',
    '6478304': 'PT LIONG PUTRA JAYA', '6478305': 'PT BUANA PINYUH AGUNG',
    '6478306': 'PT ANJUNGAN SENA PUTRA', '6478307': 'PT KILANG INDO',
    '6478308': 'PT LIAN PETRO KHATULISTIWA', '6478311': 'PT MITRA INDAH LESTARI OIL PRATAMA',
    '6478312': 'PT DUTA DAYA KALIMANTAN', '6478313': 'PT PENGKANG LESTARI',
    '6478314': 'PT SURYA PRIMA KENCANA', '6478315': 'PT SURYA INTERNUSA',
    '6478317': 'PT BUMI INDAH KHATULISTIWA', '6478318': 'PT KURNIA SAROTA ABADI',
    '6478321': 'PT BERKAH TOHA', '6478324': 'PT LAMBOY KARYA BARU',
    '6478326': 'PT BORNEO ENERGI AGUNG', '6478405': 'PT BUANA PINYUH AGUNG',
    '6478408': 'PT SEKURA JAYA', '6478501': 'PT BANGUN LINTAS NUSA',
    '6478502': 'PT SANGGAU PERMAI MINERAL', '6478503': 'PT PRIMA TAYAN LESTARI',
    '6478505': 'PT CAHAYA KEMAS PERSADA', '6478506': 'PT MAKMUR JIWANA MANDIRI',
    '6478507': 'PT JOHNDITOR', '6478508': 'PT TAYAN MITRA USAHA',
    '6478509': 'PT BALAI INDAH', '6478510': 'PT SUMBER DANA FINANCE',
    '6478511': 'PT KEPONG NIAGA', '6478512': 'PT ZULHIDJ MANDIRI',
    '6478513': 'PT TANARA KAYAN LESTARI', '6478514': 'PT TUNAS MEKAR PETRO SEJATI',
    '6478516': 'PT YOHANA CITRA MANDIRI', '6478524': 'PT TUNAS MEKAR PETRO SEJATI',
    '6478601': 'PT MELAWI TIMUR', '6478606': 'PT WILSA',
    '6478607': 'PT FIKRI DARMAWAN', '6478608': 'PT MELAWI JAYA ABADI',
    '6478609': 'PT GAUTAMA KANDAGA', '6478611': 'PT JULINDRA RAYA',
    '6478612': 'PT ANEKA CIPTA LESTARI', '6478613': 'PT DWI BERKAT SEJAHTERA',
    '6478614': 'PT AMIN PERKASA', '6478615': 'PT GAUTAMA KANDAGA',
    '6478616': 'PT CAHAYA INDAH SUBEKTI', '6478619': 'PT CAHAYA INDAH TEBELIAN',
    '6478620': 'PT BERKAH KEDAMIN JAYA', '6478705': 'PT PRATAMA BOYAN MANDIRI',
    '6478807': 'PT TRI ASTAMI PETROLIUM', '6478814': 'PT AIMAS SEJAHTERA ABADI',
    '6478816': 'PT PUTERA MANDIRI BORNEO', '6478817': 'PT PUTERA MANDIRI BORNEO',
    '6478819': 'PT LOC ANDALAN PETROLIUM', '6479101': 'PT MITRALISINDO',
    '6479102': 'PT SINGKAWANG KUMALA INDAH', '64791022': 'PT SINAR SAMUDRA SUKSES',
    '6479105': 'PT LAWIRA', '6479109': 'PT LAWIRA',
    '6479111': 'PT JO BUMI EMAS', '6479112': 'PT HARAPAN TEBAS MANDIRI',
    '6479113': 'PT BUKIT RAYA INDAH', '6479114': 'PT DHARMA MIGAS CEMERLANG',
    '6479115': 'PT RATU SEPUDAK JAYA', '6479116': 'PT CIPTASERASI KONSULINDO',
    '6479117': 'PT SAMALANTAN BATUAH', '6479118': 'PT VERONIKA MANDIRI',
    '6479119': 'PT GUNUNG JAYA', '6479120': 'PT IMMANUEL JAYA',
    '6479124': 'PT BINTANG JASA TRANSKAL', '64793001': 'PT PANCUR KALAWAR',
    '64793004': 'PT PANCUR KALAWAR', '6479304': 'PT MANDIRI SEWINDU SEJAHTERA',
    '6479305': 'PT BRUNAI', '6479306': 'PT DUTA DAYA KALIMANTAN',
    '6479307': 'PT MITRA KITA LANDAK', '6479407': 'PT PARADES',
    '6479408': 'PT SEKURA JAYA', '6479411': 'PT DENASHURYA',
    '64795003': 'PT SINAR PERKASA SEKADAU', '6479502': 'PT SINAR ABADI SEKADAU',
    '6479503': 'PT PETRO SEKADAU SEJAHTERA', '65767003': 'PT GELORA JUMHANA',
    '6578302': 'PT YUSUF PETRO ENERGI', '6578303': 'PT BUMI INDAH KHATULISTIWA',
    '6578312': 'PT DUTA DAYA KALIMANTAN', '657870003': 'PT GELORA JUMHANA',
    '65787002': 'PT UNCAK KAPUAS MANDIRI', '65787003': 'PT GELORA JUMHANA',
    '6579301': 'PT KELUARGA BESAR TARIGAS', '6579306': 'PT DUTA DAYA KALIMANTAN',
    '65794001': 'PT PALOH BERSATU JAYA ABADI', '65794002': 'PT SAJINGAN MELANCAR ABADI',
    '6678609': 'PT PRIMA JASA PREMIUM', '66787001': 'PT ERNA CITRA ABADI',
    '6678812': 'PT ANDALAN PAWAN SAHABAT PERKASA', '66793001': 'PT GAS JAYA LANDAK',
    '6679302': 'PT LANDAK JAYA PERSADA', '6679304': 'PT KALINDO MUARA BEMAYA',
    '6679305': 'PT PABAYO KARYA MANDIRI', '6679401': 'PT CARE MIGAS MAKMUR',
    '6679402': 'PT SINAR KARYA TEKARANG', '6679603': 'PT AMORA JAYA NIAGA',
    '6878313': 'PT PENGKANG LESTARI', '68788005': 'PT CARE MIGAS MAKMUR',
    '68794002': 'PT KOPERASI SAMUDERA JAYA', '7478505': 'PT CAHAYA KEMAS PERSADA',
    # Fix: merge 64781111 into 6478111
    '64781111': 'PT SOLA PETRA ABADI',
}

def download_photo(file_id, output_path):
    """Download photo from Google Drive - handle virus scan warning"""
    # Try direct download first
    result = subprocess.run([
        'curl', '-sL', '-o', output_path,
        f'https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t'
    ], capture_output=True, timeout=30)
    
    if result.returncode != 0 or not os.path.exists(output_path):
        return False
    
    size = os.path.getsize(output_path)
    if size < 2000:
        os.remove(output_path)
        return False
    
    # Check if it's actually an image (not HTML error page)
    with open(output_path, 'rb') as f:
        header = f.read(3)
    if header not in (b'\xff\xd8\xff', b'\x89PN', b'GIF8'):
        os.remove(output_path)
        return False
    
    return True

# ============================================================
# TAHAP 1: SPBU + USER
# ============================================================
def tahap1():
    print('\n' + '='*60)
    print('TAHAP 1: Import SPBU & Buat Akun')
    print('='*60)
    
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['Form Responses 1']
    
    # Extract unique SPBU dengan nama yang sudah dibenerin
    spbu_dict = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        ts, nama_pt, nomor_spbu, kota, *rest = row
        nomor = str(int(nomor_spbu)) if isinstance(nomor_spbu, (int, float)) else str(nomor_spbu or '').strip()
        kota = str(kota or '').strip().title()
        # Merge 64781111 -> 6478111
        if nomor == '64781111':
            nomor = '6478111'
        if nomor and nomor not in spbu_dict:
            nama = NAMA_BENAR.get(nomor, f'SPBU {nomor}')
            spbu_dict[nomor] = {'nama_pt': nama, 'kota': kota}
    
    print(f'SPBU unik: {len(spbu_dict)}')
    
    with app.app_context():
        created_spbu = 0
        created_user = 0
        skipped = 0
        
        for nomor, data in spbu_dict.items():
            if SPBU.query.filter_by(nomor_spbu=nomor).first():
                skipped += 1
                continue
            
            spbu = SPBU(
                nomor_spbu=nomor,
                nama_pt=data['nama_pt'],
                kota=data['kota'],
                alamat=f'SPBU {data["nama_pt"]}, {data["kota"]}',
            )
            db.session.add(spbu)
            db.session.flush()
            
            # Create user
            username = f'spbu_{nomor}'
            password = f'spbu{nomor[-4:]}'
            user = User(
                username=username, nama_lengkap=data['nama_pt'][:120],
                role='spbu', id_spbu=spbu.id,
            )
            user.set_password(password)
            db.session.add(user)
            
            created_spbu += 1
            if created_spbu % 50 == 0:
                db.session.commit()
        
        db.session.commit()
        print(f'✅ SPBU: {created_spbu} baru, {skipped} sudah ada')
        print(f'✅ User: {User.query.filter_by(role="spbu").count()} akun SPBU')

# ============================================================
# TAHAP 2: LAPORAN
# ============================================================
def tahap2():
    print('\n' + '='*60)
    print('TAHAP 2: Import Laporan')
    print('='*60)
    
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['Form Responses 1']
    
    with app.app_context():
        # SPBU cache
        spbu_cache = {}
        for s in SPBU.query.all():
            spbu_cache[s.nomor_spbu] = s
        
        admin = User.query.filter_by(role='administrator').first()
        
        total = ws.max_row - 1
        created = 0
        skipped = 0
        errors = []
        
        # Track existing (spbu_id, tanggal, bbm) to avoid dups
        existing_set = set()
        for l in Laporan.query.all():
            existing_set.add((l.id_spbu, str(l.tanggal), str(l.jenis_bbm)))
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            ts, nama_pt, nomor_spbu, kota, *rest = row
            nomor = str(int(nomor_spbu)) if isinstance(nomor_spbu, (int, float)) else str(nomor_spbu or '').strip()
            if isinstance(ts, datetime):
                tanggal = ts.date()
            else:
                continue
            
            jenis_bbm_str = str(rest[-1] if rest else '').strip()
            jenis_list = [b.strip() for b in jenis_bbm_str.split(',') if b.strip()] or ['Biosolar']
            
            spbu = spbu_cache.get(nomor)
            if not spbu:
                continue
            
            key = (spbu.id, str(tanggal), str(jenis_list))
            if key in existing_set:
                skipped += 1
                continue
            
            try:
                laporan = Laporan(
                    id_spbu=spbu.id, id_user=admin.id if admin else 1,
                    tanggal=tanggal, jenis_bbm=jenis_list,
                    status='submitted', catatan='',
                )
                db.session.add(laporan)
                existing_set.add(key)
                created += 1
                
                if created % 200 == 0:
                    db.session.commit()
                    print(f'  Progress: {created} laporan...')
            except Exception as e:
                errors.append(str(e)[:80])
        
        db.session.commit()
        print(f'✅ Laporan: {created} baru, {skipped} skip')
        if errors:
            print(f'❌ Error: {len(errors)}')
            for e in errors[:3]:
                print(f'  - {e}')

# ============================================================
# TAHAP 3: FOTO
# ============================================================
def tahap3():
    print('\n' + '='*60)
    print('TAHAP 3: Download Foto')
    print('='*60)
    t0 = time.time()
    
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['Form Responses 1']
    categories = ['pembongkaran', 'spp', 'dipping', 'atg']
    
    with app.app_context():
        spbu_map = {s.nomor_spbu: s for s in SPBU.query.all()}
        admin = User.query.filter_by(role='administrator').first()
        
        success = 0
        failed = 0
        skipped = 0
        total_links = 0
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            ts, nama_pt, nomor_spbu, kota, *rest = row
            nomor = str(int(nomor_spbu)) if isinstance(nomor_spbu, (int, float)) else str(nomor_spbu or '').strip()
            tanggal = ts.date() if isinstance(ts, datetime) else datetime.now().date()
            
            spbu = spbu_map.get(nomor)
            if not spbu:
                continue
            
            laporan = Laporan.query.filter_by(id_spbu=spbu.id, tanggal=tanggal).first()
            if not laporan:
                continue
            
            for ci, cat in enumerate(categories):
                url_str = str(rest[ci] or '') if ci < len(rest) else ''
                if not url_str:
                    continue
                
                # Split multiple URLs
                urls = [u.strip() for u in re.split(r'[,\n\r]+', url_str) if u.strip()]
                for single_url in urls:
                    total_links += 1
                    file_id = re.search(r'id=([a-zA-Z0-9_-]+)', single_url)
                    if not file_id:
                        failed += 1
                        continue
                    file_id = file_id.group(1)
                    
                    # Check if already uploaded
                    existing = Upload.query.filter(
                        Upload.id_laporan == laporan.id,
                        Upload.path.contains(file_id[:12])
                    ).first()
                    if existing:
                        skipped += 1
                        continue
                    
                    # Download
                    temp_dir = os.path.join(UPLOAD_DIR, str(laporan.id))
                    os.makedirs(temp_dir, exist_ok=True)
                    fname = f'{cat}_{file_id[:10]}.jpg'
                    temp_path = os.path.join(temp_dir, f'temp_{fname}')
                    out_path = os.path.join(temp_dir, fname)
                    thumb_path = os.path.join(temp_dir, f'thumb_{fname}')
                    
                    if not os.path.exists(out_path) or os.path.getsize(out_path) < 1000:
                        ok = download_photo(file_id, temp_path)
                        if not ok:
                            failed += 1
                            continue
                        # Compress image
                        compress_image(temp_path, out_path)
                        # Generate thumbnail
                        try:
                            generate_thumbnail(temp_path, thumb_path)
                        except:
                            pass
                        # Remove temp file
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                    elif not os.path.exists(thumb_path):
                        try:
                            generate_thumbnail(out_path, thumb_path)
                        except:
                            pass
                    
                    # Create upload record
                    try:
                        upload = Upload(
                            id_laporan=laporan.id, kategori=cat,
                            filename=fname, path=f'{laporan.id}/{fname}',
                        )
                        db.session.add(upload)
                        db.session.commit()
                        success += 1
                    except:
                        db.session.rollback()
                        failed += 1
            
            if idx % 100 == 0:
                elapsed = time.time() - t0
                rate = success / elapsed if elapsed > 0 else 0
                print(f'  Row {idx}: OK={success} Fail={failed} Skip={skipped} Rate={rate:.1f}/s')
        
        print(f'\n✅ Foto: {success} sukses, {failed} gagal, {skipped} skip')
        print(f'Total link diproses: {total_links}')

# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    t0 = time.time()
    
    tahap1()
    tahap2()
    tahap3()
    
    elapsed = time.time() - t0
    print(f'\n⏱ Total: {elapsed/60:.1f} menit')
