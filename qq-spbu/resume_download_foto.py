#!/usr/bin/env python3
"""Resume downloading missing photos from Excel to QQ SPBU."""
import time, os, sys, re, subprocess
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.environ['FLASK_ENV'] = 'production'
os.environ['DATABASE_URL'] = 'postgresql://qqspbu:qqspbu2026@localhost:5432/qqspbu'
os.environ['SECRET_KEY'] = 'qq-spbu-secret-key-2026-kalbar'

sys.path.insert(0, BASE_DIR)
import openpyxl
from models import db, SPBU, Laporan, Upload
from utils import compress_image, generate_thumbnail
from app import app

EXCEL = '/root/.hermes/cache/documents/doc_82239650f0f4_Laporan_Pembongkaran_dan_Penerimaan_BBM_di_SPBU_MTD_14_Juli_2026.xlsx'
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
categories = ['pembongkaran', 'spp', 'dipping', 'atg']

def download_photo(file_id, output_path):
    """Download photo from Google Drive."""
    result = subprocess.run([
        'curl', '-sL', '--max-time', '25', '-o', output_path,
        f'https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t'
    ], capture_output=True, timeout=30)
    if result.returncode != 0 or not os.path.exists(output_path):
        return False
    size = os.path.getsize(output_path)
    if size < 2000:
        os.remove(output_path)
        return False
    with open(output_path, 'rb') as f:
        header = f.read(3)
    if header not in (b'\xff\xd8\xff', b'\x89PN', b'GIF8'):
        os.remove(output_path)
        return False
    return True

print('TAHAP 3: Download Foto (resume)')
print()

with app.app_context():
    spbu_map = {s.nomor_spbu: s for s in SPBU.query.all()}
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['Form Responses 1']

    success = 0
    failed = 0
    skipped = 0
    total_links = 0
    total_rows = ws.max_row - 1
    t_start = time.time()
    last_progress = time.time()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        ts, nama_pt, nomor_spbu, kota, *rest = row
        nomor = str(int(nomor_spbu)) if isinstance(nomor_spbu, (int, float)) else str(nomor_spbu or '').strip()
        try:
            tanggal = ts.date() if isinstance(ts, datetime) else None
        except:
            tanggal = None
        if not tanggal:
            continue

        spbu = spbu_map.get(nomor)
        if not spbu:
            continue

        laporan = Laporan.query.filter_by(id_spbu=spbu.id, tanggal=tanggal).first()
        if not laporan:
            continue

        for ci, cat in enumerate(categories):
            url_str = str(rest[ci] or '') if ci < len(rest) else ''
            if not url_str:
                continue
            urls = [u.strip() for u in re.split(r'[,\n\r]+', url_str) if u.strip() and 'drive' in u.lower()]

            for single_url in urls:
                total_links += 1
                file_id_match = re.search(r'id=([a-zA-Z0-9_-]+)', single_url)
                if not file_id_match:
                    failed += 1
                    continue
                file_id = file_id_match.group(1)

                # Check if already exists
                existing = Upload.query.filter(
                    Upload.id_laporan == laporan.id,
                    Upload.path.contains(file_id[:10])
                ).first()
                if existing:
                    skipped += 1
                    continue

                temp_dir = os.path.join(UPLOAD_DIR, str(laporan.id))
                os.makedirs(temp_dir, exist_ok=True)
                fname = f'{cat}_{file_id[:10]}.jpg'
                temp_path = os.path.join(temp_dir, f'temp_{fname}')
                out_path = os.path.join(temp_dir, fname)
                thumb_path = os.path.join(temp_dir, f'thumb_{fname}')

                if not os.path.exists(out_path) or os.path.getsize(out_path) < 1000:
                    try:
                        ok = download_photo(file_id, temp_path)
                    except subprocess.TimeoutExpired:
                        failed += 1
                        if failed <= 20 or failed % 50 == 0:
                            print(f'  ⚠️ TIMEOUT Row {idx} {nomor} {cat}')
                        continue
                    except Exception as e:
                        failed += 1
                        if failed <= 20 or failed % 50 == 0:
                            print(f'  ⚠️ ERROR Row {idx} {nomor} {cat}: {str(e)[:60]}')
                        continue
                    if not ok:
                        failed += 1
                        continue
                    try:
                        compress_image(temp_path, out_path)
                        generate_thumbnail(temp_path, thumb_path)
                    except:
                        pass
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                elif not os.path.exists(thumb_path):
                    try:
                        generate_thumbnail(out_path, thumb_path)
                    except:
                        pass

                upload = Upload(
                    id_laporan=laporan.id, kategori=cat,
                    filename=fname, path=f'{laporan.id}/{fname}',
                )
                db.session.add(upload)
                try:
                    db.session.commit()
                    success += 1
                except Exception as e:
                    db.session.rollback()
                    failed += 1

        # Progress every 100 rows OR every 30 seconds
        now = time.time()
        if idx % 100 == 0 or (now - last_progress > 30):
            elapsed = now - t_start
            rate = success / elapsed if elapsed > 0 else 0
            remain = total_rows - idx
            row_rate = idx / elapsed if elapsed > 0 else 0
            eta = remain / row_rate if row_rate > 0 else 0
            print(f'  Row {idx:4d}/{total_rows}: OK={success:5d} Fail={failed:4d} Skip={skipped:4d} Rate={rate:.1f}/s ETA={eta:.0f}s')
            last_progress = now
            db.session.commit()

    elapsed = time.time() - t_start
    print(f'\n{"="*55}')
    print(f'✅ SELESAI!')
    print(f'  OK={success}  Fail={failed}  Skip={skipped}  Total links={total_links}')
    print(f'  ⏱ {elapsed:.0f}s ({elapsed/60:.1f} menit)')
    print(f'  Rata-rata: {success/elapsed:.1f} foto/detik')
