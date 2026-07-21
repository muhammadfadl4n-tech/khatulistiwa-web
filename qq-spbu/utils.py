"""
Utility functions untuk file upload, image compression, dan helpers.
"""

import os
import uuid
import io
from datetime import datetime
from functools import wraps
from flask import abort, current_app
from werkzeug.utils import secure_filename
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def allowed_file(filename):
    """Check apakah file extension diizinkan."""
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in current_app.config['ALLOWED_EXTENSIONS']


def is_image(filename):
    """Check apakah file adalah image (bukan PDF)."""
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in current_app.config['ALLOWED_IMAGE_EXTENSIONS']


def generate_unique_filename(original_filename):
    """Generate unique filename dengan timestamp dan UUID."""
    ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_id = uuid.uuid4().hex[:8]
    secure_name = secure_filename(original_filename.rsplit('.', 1)[0])
    
    if ext:
        return f"{timestamp}_{unique_id}_{secure_name}.{ext}"
    return f"{timestamp}_{unique_id}_{secure_name}"


def get_upload_subfolder(laporan_id):
    """Generate subfolder path berdasarkan tanggal dan laporan ID."""
    now = datetime.now()
    return os.path.join(str(now.year), f"{now.month:02d}", str(laporan_id))


def compress_image(input_path, output_path, max_dim=None, quality=None):
    """
    Compress dan resize gambar.
    
    Args:
        input_path: Path ke file gambar asli
        output_path: Path untuk menyimpan hasil kompresi
        max_dim: Maximum dimension (longest side) dalam pixels
        quality: JPEG quality (1-100)
    
    Returns:
        dict: Informasi hasil kompresi (ukuran, dimensi)
    """
    if max_dim is None:
        max_dim = current_app.config.get('IMAGE_MAX_DIMENSION', 1920)
    if quality is None:
        quality = current_app.config.get('IMAGE_QUALITY', 80)
    
    # Buka gambar
    img = Image.open(input_path)
    
    # Strip EXIF data dan convert ke RGB jika perlu
    if img.mode in ('RGBA', 'P', 'LA'):
        # Convert transparent images ke RGB dengan background putih
        background = Image.new('RGB', img.size, (255, 255, 255))
        if img.mode == 'P':
            img = img.convert('RGBA')
        if 'A' in img.mode:
            background.paste(img, mask=img.split()[-1])
        else:
            background.paste(img)
        img = background
    elif img.mode != 'RGB':
        img = img.convert('RGB')
    
    # Resize jika melebihi max dimension
    original_size = img.size
    if max(img.size) > max_dim:
        ratio = max_dim / max(img.size)
        new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
        img = img.resize(new_size, Image.Resampling.LANCZOS)
    
    # Simpan dengan kompresi
    img.save(output_path, 'JPEG', quality=quality, optimize=True)
    
    # Get file size
    output_size = os.path.getsize(output_path)
    
    return {
        'original_size': original_size,
        'new_size': img.size,
        'file_size': output_size,
    }


def generate_thumbnail(input_path, output_path, size=None):
    """
    Generate thumbnail dari gambar.
    
    Args:
        input_path: Path ke file gambar
        output_path: Path untuk menyimpan thumbnail
        size: Ukuran thumbnail (width, height)
    
    Returns:
        dict: Informasi thumbnail
    """
    if size is None:
        size = current_app.config.get('THUMBNAIL_SIZE', 320)
    
    img = Image.open(input_path)
    
    # Convert ke RGB jika perlu
    if img.mode != 'RGB':
        if img.mode in ('RGBA', 'P', 'LA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            if 'A' in img.mode:
                background.paste(img, mask=img.split()[-1])
            else:
                background.paste(img)
            img = background
        else:
            img = img.convert('RGB')
    
    # Generate thumbnail
    img.thumbnail((size, size), Image.Resampling.LANCZOS)
    img.save(output_path, 'JPEG', quality=75, optimize=True)
    
    return {
        'size': img.size,
        'file_size': os.path.getsize(output_path),
    }


def save_upload(file, laporan_id, kategori):
    """
    Simpan file upload dengan kompresi (untuk image) dan generate thumbnail.
    
    Args:
        file: File object dari request.files
        laporan_id: ID laporan
        kategori: Kategori upload (pembongkaran, spp, dipping, atg)
    
    Returns:
        dict: Informasi file yang disimpan (filename, path, thumbnail_path)
    """
    if not file or file.filename == '':
        return None
    
    if not allowed_file(file.filename):
        raise ValueError(f"File type tidak diizinkan: {file.filename}")
    
    # Generate unique filename
    original_filename = secure_filename(file.filename)
    unique_filename = generate_unique_filename(original_filename)
    
    # Generate subfolder
    subfolder = get_upload_subfolder(laporan_id)
    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], subfolder)
    
    # Create directory jika belum ada
    os.makedirs(upload_dir, exist_ok=True)
    
    # Save file ke temporary path
    temp_path = os.path.join(upload_dir, f"temp_{unique_filename}")
    file.save(temp_path)
    
    # Use forward slash for URL compatibility (Windows uses backslash)
    url_path = f"{subfolder}/{unique_filename}".replace('\\', '/')
    
    result = {
        'filename': original_filename,
        'path': url_path,
        'thumbnail_path': None,
    }
    
    # Compress jika image
    if is_image(original_filename):
        output_path = os.path.join(upload_dir, unique_filename)
        compress_result = compress_image(temp_path, output_path)
        
        # Generate thumbnail
        thumbnail_filename = f"thumb_{unique_filename}"
        thumbnail_path = os.path.join(upload_dir, thumbnail_filename)
        generate_thumbnail(temp_path, thumbnail_path)
        
        url_thumbnail_path = f"{subfolder}/{thumbnail_filename}".replace('\\', '/')
        
        result['path'] = url_path
        result['thumbnail_path'] = url_thumbnail_path
        result['compression'] = compress_result
        
        # Hapus temporary file
        os.remove(temp_path)
    else:
        # Untuk PDF, langsung rename
        final_path = os.path.join(upload_dir, unique_filename)
        os.rename(temp_path, final_path)
        result['path'] = url_path
    
    return result


def role_required(*roles):
    """
    Decorator untuk membatasi akses berdasarkan role.
    
    Usage:
        @role_required('administrator', 'pertamina')
        def admin_only_route():
            ...
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            from flask_login import current_user
            if not current_user.is_authenticated:
                abort(401)
            if current_user.role not in roles:
                abort(403)
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def spbu_owner_required(f):
    """
    Decorator untuk memastikan SPBU hanya bisa akses laporan miliknya sendiri.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        from flask_login import current_user
        from models import Laporan
        
        if not current_user.is_authenticated:
            abort(401)
        
        # Administrator dan Pertamina bisa akses semua
        if current_user.role in ('administrator', 'pertamina'):
            return f(*args, **kwargs)
        
        # SPBU hanya bisa akses laporan miliknya
        if current_user.role == 'spbu':
            laporan_id = kwargs.get('id') or kwargs.get('laporan_id')
            if laporan_id:
                laporan = Laporan.query.get(laporan_id)
                if laporan and laporan.id_spbu != current_user.id_spbu:
                    abort(403)
        
        return f(*args, **kwargs)
    return decorated_function


# ============================================================
# EXCEL EXPORT/IMPORT UTILITIES
# ============================================================

def create_excel_styles():
    """Create common styles for Excel exports."""
    return {
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'header_fill': PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
    }


def export_spbu_to_excel(spbu_list):
    """
    Export daftar SPBU ke Excel.
    
    Args:
        spbu_list: List of SPBU objects
    
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data SPBU"
    
    styles = create_excel_styles()
    
    # Headers
    headers = ['No', 'Nomor SPBU', 'Nama PT/CV', 'Kota', 'Alamat', 'Latitude', 'Longitude', 'Tanggal Daftar']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Data rows
    for idx, spbu in enumerate(spbu_list, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=spbu.nomor_spbu)
        ws.cell(row=idx + 1, column=3, value=spbu.nama_pt)
        ws.cell(row=idx + 1, column=4, value=spbu.kota)
        ws.cell(row=idx + 1, column=5, value=spbu.alamat)
        ws.cell(row=idx + 1, column=6, value=spbu.lat)
        ws.cell(row=idx + 1, column=7, value=spbu.lng)
        ws.cell(row=idx + 1, column=8, value=spbu.created_at.strftime('%Y-%m-%d') if spbu.created_at else '')
        
        # Apply borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=idx + 1, column=col).border = styles['border']
    
    # Auto-fit column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_users_to_excel(user_list):
    """
    Export daftar User ke Excel.
    
    Args:
        user_list: List of User objects
    
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data User"
    
    styles = create_excel_styles()
    
    # Headers
    headers = ['No', 'Username', 'Nama Lengkap', 'Email', 'Role', 'SPBU', 'Status', 'Terdaftar']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Data rows
    for idx, user in enumerate(user_list, 1):
        spbu_info = ''
        if user.role == 'spbu' and user.spbu:
            spbu_info = f"{user.spbu.nomor_spbu} - {user.spbu.nama_pt}"
        
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=user.username)
        ws.cell(row=idx + 1, column=3, value=user.nama_lengkap)
        ws.cell(row=idx + 1, column=4, value=user.email or '')
        ws.cell(row=idx + 1, column=5, value=user.role.capitalize())
        ws.cell(row=idx + 1, column=6, value=spbu_info)
        ws.cell(row=idx + 1, column=7, value='Aktif' if user.is_active else 'Nonaktif')
        ws.cell(row=idx + 1, column=8, value=user.created_at.strftime('%Y-%m-%d') if user.created_at else '')
        
        # Apply borders
        for col in range(1, len(headers) + 1):
            ws.cell(row=idx + 1, column=col).border = styles['border']
    
    # Auto-fit column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_spbu_template():
    """
    Generate template Excel kosong untuk import SPBU.
    
    Returns:
        BytesIO: Excel template in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import SPBU"
    
    styles = create_excel_styles()
    
    # Headers
    headers = ['Nomor SPBU', 'Nama PT/CV', 'Kota', 'Alamat', 'Latitude', 'Longitude']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Example row
    example_data = ['64.751.99', 'PT. Contoh Sejahtera', 'Pontianak', 'Jl. Contoh No. 123', '-0.0263', '109.3425']
    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        cell.border = styles['border']
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_user_template():
    """
    Generate template Excel kosong untuk import User.
    
    Returns:
        BytesIO: Excel template in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import User"
    
    styles = create_excel_styles()
    
    # Headers
    headers = ['Username', 'Password', 'Nama Lengkap', 'Email', 'Role', 'Nomor SPBU']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_alignment']
        cell.border = styles['border']
    
    # Example rows
    example_data = [
        ['spbu99', 'password123', 'Operator Contoh', 'contoh@email.com', 'spbu', '64.751.99'],
        ['pertamina99', 'password123', 'Staff Pertamina', 'pertamina@email.com', 'pertamina', ''],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            cell.border = styles['border']
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_spbu_from_excel(file_stream):
    """
    Import SPBU data dari Excel file.
    
    Args:
        file_stream: File stream dari uploaded file
    
    Returns:
        tuple: (success_count, error_list)
    """
    from openpyxl import load_workbook
    from models import db, SPBU
    
    wb = load_workbook(file_stream)
    ws = wb.active
    
    success_count = 0
    error_list = []
    
    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:  # Skip empty rows
            continue
        
        nomor_spbu = str(row[0]).strip() if row[0] else ''
        nama_pt = str(row[1]).strip() if row[1] else ''
        kota = str(row[2]).strip() if row[2] else ''
        alamat = str(row[3]).strip() if row[3] else ''
        lat = row[4] if row[4] else None
        lng = row[5] if row[5] else None
        
        # Validation
        if not nomor_spbu or not nama_pt or not kota:
            error_list.append(f"Baris {row_idx}: Nomor SPBU, Nama PT, dan Kota wajib diisi")
            continue
        
        # Check duplicate
        existing = SPBU.query.filter_by(nomor_spbu=nomor_spbu).first()
        if existing:
            error_list.append(f"Baris {row_idx}: Nomor SPBU '{nomor_spbu}' sudah terdaftar")
            continue
        
        # Convert lat/lng
        try:
            lat = float(lat) if lat else None
            lng = float(lng) if lng else None
        except (ValueError, TypeError):
            lat = None
            lng = None
        
        # Create new SPBU
        new_spbu = SPBU(
            nomor_spbu=nomor_spbu,
            nama_pt=nama_pt,
            kota=kota,
            alamat=alamat,
            lat=lat,
            lng=lng
        )
        db.session.add(new_spbu)
        success_count += 1
    
    if success_count > 0:
        db.session.commit()
    
    return success_count, error_list


def import_users_from_excel(file_stream):
    """
    Import User data dari Excel file.
    
    Args:
        file_stream: File stream dari uploaded file
    
    Returns:
        tuple: (success_count, error_list)
    """
    from openpyxl import load_workbook
    from models import db, User, SPBU
    
    wb = load_workbook(file_stream)
    ws = wb.active
    
    success_count = 0
    error_list = []
    
    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:  # Skip empty rows
            continue
        
        username = str(row[0]).strip() if row[0] else ''
        password = str(row[1]).strip() if row[1] else ''
        nama_lengkap = str(row[2]).strip() if row[2] else ''
        email = str(row[3]).strip() if row[3] else ''
        role = str(row[4]).strip().lower() if row[4] else 'spbu'
        nomor_spbu = str(row[5]).strip() if row[5] else ''
        
        # Validation
        if not username or not password or not nama_lengkap:
            error_list.append(f"Baris {row_idx}: Username, Password, dan Nama Lengkap wajib diisi")
            continue
        
        if role not in ['spbu', 'pertamina', 'administrator']:
            error_list.append(f"Baris {row_idx}: Role tidak valid '{role}'. Gunakan: spbu, pertamina, atau administrator")
            continue
        
        # Check duplicate username
        existing = User.query.filter_by(username=username).first()
        if existing:
            error_list.append(f"Baris {row_idx}: Username '{username}' sudah terdaftar")
            continue
        
        # Find SPBU if role is spbu
        id_spbu = None
        if role == 'spbu' and nomor_spbu:
            spbu = SPBU.query.filter_by(nomor_spbu=nomor_spbu).first()
            if not spbu:
                error_list.append(f"Baris {row_idx}: Nomor SPBU '{nomor_spbu}' tidak ditemukan")
                continue
            id_spbu = spbu.id
        
        # Create new User
        new_user = User(
            username=username,
            nama_lengkap=nama_lengkap,
            email=email if email else None,
            role=role,
            id_spbu=id_spbu,
            is_active=True
        )
        new_user.set_password(password)
        db.session.add(new_user)
        success_count += 1
    
    if success_count > 0:
        db.session.commit()
    
    return success_count, error_list
